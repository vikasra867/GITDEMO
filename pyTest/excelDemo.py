import openpyxl

book = openpyxl.load_workbook("D:\\SeleniumPython\\excelDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=1)
print(cell.value)

sheet.cell(row=6, column=6).value = "Col6Row6"
print(sheet['A5'].value)

for i in range(2, sheet.max_column + 1):
    for j in range(1, sheet.max_row + 1):
        print(sheet.cell(row=j, column=i).value)
        Dict = [sheet.cell(row=1, column=i).value] = sheet.cell(row=j, column=i).value
